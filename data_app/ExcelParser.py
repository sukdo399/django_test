import os
import logging

from openpyxl import load_workbook
from openpyxl import Workbook
# from openpyxl.comments import Comment
# from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Alignment

logger = logging.getLogger('logger')


class ExcelParser(object):

    def get_data(file_path):
        logger.debug("file_path: %s" % file_path)
        try:
            wb = load_workbook(file_path)
        except:
            logger.error("it's not a excel file")
            raise NameError("it's not a excel file")

        logger.debug("sheet_names: %s" % wb.get_sheet_names())

        ws = wb.get_sheet_by_name('Entropy')

        data = []
        row_data = []
        for row in ws.iter_rows():
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
            row_data = []

        return data
